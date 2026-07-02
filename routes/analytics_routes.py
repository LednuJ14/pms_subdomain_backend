from flask import Blueprint, jsonify, current_app, request, send_file, make_response
from flask_jwt_extended import jwt_required, get_jwt_identity, get_jwt
from sqlalchemy import func, extract, or_, text
from datetime import datetime, timedelta, date, timezone
from decimal import Decimal
import re
import csv
import io

from app import db
from models.user import User, UserRole
from models.property import Property, Unit, UnitStatus
from models.tenant import Tenant
from models.staff import Staff, EmploymentStatus
from models.bill import Bill, BillStatus, Payment, PaymentStatus
from models.request import MaintenanceRequest, RequestStatus
from models.announcement import Announcement
from models.task import Task, TaskStatus

# Try to import reportlab for PDF generation
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    import logging
    logging.warning("reportlab not available. PDF reports will not work.")

# Try to import openpyxl for Excel generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    import logging
    logging.warning("openpyxl not available. Excel reports will not work.")

# Try to import TenantUnit, but handle if it doesn't exist
try:
    from models.tenant import TenantUnit
    TENANT_UNIT_AVAILABLE = True
except ImportError:
    TENANT_UNIT_AVAILABLE = False
    TenantUnit = None

analytics_bp = Blueprint('analytics', __name__)

from models.user import User

def is_super_admin(user_id):
    if not user_id: return False
    user = User.query.get(user_id)
    return user and getattr(user, 'role', '') == 'ADMIN'


def get_property_id_from_request(data=None):
    """
    Try to get property_id from request.
    Checks request body, query parameter, header, subdomain, or Origin header.
    Returns None if not found.
    """
    try:
        # 1. Check query parameters
        property_id_raw = request.args.get('property_id')
        subdomain_param = request.args.get('subdomain')
        
        if property_id_raw:
            try:
                return int(property_id_raw)
            except ValueError:
                # It's a string (e.g. 'horizon'), treat it as a subdomain
                if not subdomain_param:
                    subdomain_param = property_id_raw
                    
        # 2. Check headers
        header_property_id = request.headers.get('X-Property-ID')
        if header_property_id:
            try:
                return int(header_property_id)
            except ValueError:
                # It's a string, treat it as a subdomain
                if not subdomain_param:
                    subdomain_param = header_property_id

        # 3. Resolve subdomain if we found one in query or header
        if subdomain_param:
            subdomain_val = str(subdomain_param).lower().strip()
            if subdomain_val != 'admin':
                try:
                    # Try exact match on portal_subdomain, title, building_name
                    match_columns = ['portal_subdomain', 'title', 'building_name']
                    for col in match_columns:
                        try:
                            property_obj = db.session.execute(text(
                                f"SELECT id FROM properties WHERE LOWER(TRIM(COALESCE({col}, ''))) = :subdomain LIMIT 1"
                            ), {'subdomain': subdomain_val}).first()
                            
                            if property_obj:
                                return property_obj[0]
                        except Exception:
                            continue
                except Exception as e:
                    current_app.logger.warning(f"Error matching subdomain {subdomain_val}: {str(e)}")

        # 4. Check JWT claims
        try:
            claims = get_jwt()
            if claims:
                property_id = claims.get('property_id')
                if property_id:
                    return int(property_id)
        except Exception:
            pass
        
        # 5. Check request body if data is provided
        if data:
            property_id = data.get('property_id')
            if property_id:
                try:
                    return int(property_id)
                except (ValueError, TypeError):
                    pass
        
        # 6. Try to extract from subdomain in Origin or Host header (Fallback)
        origin = request.headers.get('Origin', '')
        host = request.headers.get('Host', '')
        
        if origin or host:
            # Extract subdomain (works for both .localhost and production .pms.vicirotechnologies.com)
            # This matches the first part of the domain before the first dot
            host_to_check = origin.replace('https://', '').replace('http://', '') if origin else host
            if host_to_check:
                parts = host_to_check.split(':')[0].split('.')
                if len(parts) > 1 and parts[0].lower() != 'localhost' and not parts[0].replace('.', '').isdigit():
                    subdomain = parts[0].lower()
                    
                    if subdomain != 'admin':
                        try:
                            match_columns = ['portal_subdomain', 'title', 'building_name']
                            for col in match_columns:
                                try:
                                    property_obj = db.session.execute(text(
                                        f"SELECT id FROM properties WHERE LOWER(TRIM(COALESCE({col}, ''))) = :subdomain LIMIT 1"
                                    ), {'subdomain': subdomain}).first()
                                    
                                    if property_obj:
                                        return property_obj[0]
                                except Exception:
                                    continue
                        except Exception:
                            pass
        
        return None
    except Exception as e:
        current_app.logger.warning(f"Error getting property_id from request: {str(e)}")
        return None

def table_exists(table_name):
    """Check if a table exists in the database."""
    try:
        result = db.session.execute(db.text(
            f"SELECT COUNT(*) FROM information_schema.TABLES "
            f"WHERE TABLE_SCHEMA = DATABASE() AND TABLE_NAME = '{table_name}'"
        ))
        return result.scalar() > 0
    except Exception:
        return False

def require_role(allowed_roles):
    """Decorator to require specific user roles."""
    def decorator(f):
        from functools import wraps
        @wraps(f)
        def decorated_function(*args, **kwargs):
            claims = get_jwt()
            user_role = claims.get('role')
            if user_role not in allowed_roles:
                return jsonify({'error': 'Insufficient permissions'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@analytics_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard_data():
    """
    Get dashboard data
    ---
    tags:
      - Analytics
    summary: Get dashboard data based on user role
    description: Get dashboard data based on user role and property context
    security:
      - Bearer: []
    responses:
      200:
        description: Dashboard data retrieved successfully
        schema:
          type: object
          properties:
            stats:
              type: object
            recent_activity:
              type: array
              items:
                type: object
      401:
        description: Unauthorized
      500:
        description: Server error
    """
    try:
        current_user_id = get_jwt_identity()
        claims = get_jwt()
        user_role = claims.get('role')
        
        # Get property_id from request (subdomain, query param, header, or JWT)
        property_id = get_property_id_from_request()
        
        # If property_id not in request, try to get from JWT token
        if not property_id:
            try:
                property_id = claims.get('property_id')
            except Exception:
                pass
        
        # CRITICAL: Do NOT auto-detect from owned properties
        # Property managers must access through the correct subdomain
        user_role_upper = str(user_role).upper() if user_role else ''
        if user_role == 'property_manager' or user_role_upper == 'ADMIN':
            if not property_id:
                # Return safe empty dashboard instead of error
                return jsonify({
                    'property_id': None,
                    'property_name': None,
                    'metrics': {
                        'total_income': 0.0,
                        'current_month': datetime.now().strftime('%B %Y'),
                        'active_tenants': 0,
                        'active_staff': 0,
                        'total_properties': 0,
                        'occupancy_rate': 0.0,
                        'outstanding_balance': 0.0
                    },
                    'properties': {
                        'total': 0,
                        'total_units': 0,
                        'occupied_units': 0,
                        'available_units': 0,
                        'occupancy_rate': 0.0
                    },
                    'sales_data': [],
                    'maintenance_requests': [],
                    'pending_tasks': [],
                    'announcements': [],
                    'message': 'No property selected. Please select a property to view analytics.'
                }), 200
            return get_manager_dashboard(property_id)
        elif user_role == 'staff':
            try:
                # Convert user_id to int if needed
                try:
                    user_id_int = int(current_user_id) if current_user_id else None
                except (ValueError, TypeError):
                    user_id_int = current_user_id
                
                return get_staff_dashboard(user_id_int, property_id)
            except Exception as staff_err:
                current_app.logger.error(f"Error calling get_staff_dashboard: {str(staff_err)}", exc_info=True)
                # Return safe default response
                return jsonify({
                    'staff_info': {
                        'id': current_user_id,
                        'name': 'Staff Member',
                        'email': ''
                    },
                    'tasks': {
                        'pending_count': 0,
                        'completed_count': 0,
                        'recent_tasks': []
                    },
                    'maintenance_requests': [],
                    'announcements': [],
                    'error': 'Failed to load dashboard data'
                }), 200
        elif user_role == 'tenant':
            return get_tenant_dashboard(current_user_id)
        else:
            return jsonify({'error': 'Invalid user role'}), 400
            
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        current_app.logger.error(f"Dashboard error: {str(e)}\n{error_trace}", exc_info=True)
        
        # Return safe default response instead of crashing
        safe_response = {
            'property_id': None,
            'property_name': None,
            'metrics': {
                'total_income': 0.0,
                'current_month': datetime.now().strftime('%B %Y'),
                'active_tenants': 0,
                'active_staff': 0,
                'total_properties': 0,
                'occupancy_rate': 0.0,
                'outstanding_balance': 0.0
            },
            'properties': {
                'total': 0,
                'total_units': 0,
                'occupied_units': 0,
                'available_units': 0,
                'occupancy_rate': 0.0
            },
            'sales_data': [],
            'maintenance_requests': [],
            'pending_tasks': [],
            'announcements': [],
            'error': 'Failed to load dashboard data',
            'error_details': str(e) if current_app.config.get('DEBUG', False) else None
        }
        
        # Return 200 with error message instead of 500 to prevent frontend crashes
        return jsonify(safe_response), 200

def get_manager_dashboard(property_id):
    """Get property manager dashboard data for a specific property."""
    try:
        current_app.logger.debug(f"Getting manager dashboard for property_id: {property_id}")
        # Verify property exists
        property_obj = Property.query.get(property_id)
        if not property_obj:
            current_app.logger.warning(f"Property {property_id} not found")
            return jsonify({'error': 'Property not found'}), 404
        
        # CRITICAL: Verify current user owns this property
        current_user_id = get_jwt_identity()
        if current_user_id:
            from models.user import User
            current_user = User.query.get(current_user_id)
            if current_user and current_user.is_property_manager():
                if property_obj.owner_id != current_user.id and not is_super_admin(current_user.id):
                    return jsonify({
                        'error': 'Access denied. You do not own this property.',
                        'code': 'PROPERTY_ACCESS_DENIED'
                    }), 403
        
        # Key Metrics - Property-specific with error handling
        try:
            # Single property context
            total_properties = 1
        except Exception as e:
            current_app.logger.warning(f"Error getting property: {str(e)}")
            total_properties = 0
        
        try:
            if table_exists('units'):
                # Only count units for this property
                total_units = Unit.query.filter_by(property_id=property_id).count()
            else:
                current_app.logger.warning("Units table does not exist")
                total_units = 0
        except Exception as e:
            current_app.logger.warning(f"Error getting total units: {str(e)}")
            total_units = 0
        
        # Calculate occupied and available units based on actual tenant-unit relationships
        # This ensures consistency with active_tenants count
        occupied_units = 0
        active_tenants = 0
        
        try:
            # Try to get occupied units and active tenants from TenantUnit relationships
            if TENANT_UNIT_AVAILABLE and TenantUnit and table_exists('tenant_units'):
                try:
                    today = date.today()
                    
                    # Get distinct units that have active tenants (occupied units)
                    occupied_unit_ids = db.session.query(Unit.id).join(TenantUnit).filter(
                        Unit.property_id == property_id,
                        TenantUnit.move_in_date.isnot(None),
                        or_(
                            TenantUnit.move_out_date.is_(None),  # Ongoing lease (no move-out date)
                            TenantUnit.move_out_date >= today   # Future move-out date
                        )
                    ).distinct().all()
                    
                    occupied_units = len(occupied_unit_ids)
                    
                    # Count active tenants
                    active_tenants = Tenant.query.join(TenantUnit).join(Unit).filter(
                        Unit.property_id == property_id,
                        TenantUnit.move_in_date.isnot(None),
                        or_(
                            TenantUnit.move_out_date.is_(None),  # Ongoing lease (no move-out date)
                            TenantUnit.move_out_date >= today   # Future move-out date
                        )
                    ).count()
                    
                    current_app.logger.debug(f"Property {property_id}: {active_tenants} active tenants in {occupied_units} occupied units")
                    
                except Exception as join_error:
                    current_app.logger.warning(f"Error joining TenantUnit: {str(join_error)}")
                    # Fallback: use unit status if TenantUnit join fails
                    try:
                        occupied_units = Unit.query.filter(
                            Unit.property_id == property_id,
                            or_(Unit.status == 'occupied', Unit.status == 'rented')
                        ).count()
                        active_tenants = db.session.query(Tenant).join(TenantUnit).join(Unit).filter(
                            Unit.property_id == property_id
                        ).count()
                        current_app.logger.debug(f"Using fallback: {active_tenants} tenants, {occupied_units} occupied units (by status)")
                    except Exception:
                        occupied_units = 0
                        active_tenants = 0
            else:
                # TenantUnit table doesn't exist - fallback to unit status
                current_app.logger.warning(f"TenantUnit not available for property {property_id}, using unit status")
                try:
                    occupied_units = Unit.query.filter(
                        Unit.property_id == property_id,
                        or_(Unit.status == 'occupied', Unit.status == 'rented')
                    ).count()
                except Exception as e:
                    current_app.logger.warning(f"Error getting occupied units by status: {str(e)}")
                    occupied_units = 0
                active_tenants = 0
        except Exception as e:
            current_app.logger.warning(f"Error getting occupied units and active tenants: {str(e)}")
            occupied_units = 0
            active_tenants = 0
        
        # Calculate available units: total - occupied
        try:
            available_units = max(0, total_units - occupied_units)
        except Exception as e:
            current_app.logger.warning(f"Error calculating available units: {str(e)}")
            available_units = 0
        
        # Count active staff for this specific property
        # Note: Staff model doesn't have employment_status column - all staff are active
        # So we just count all staff members for this property
        try:
            if table_exists('staff'):
                active_staff = Staff.query.filter_by(property_id=property_id).count()
                current_app.logger.debug(f"Property {property_id}: Found {active_staff} staff members")
            else:
                current_app.logger.warning("Staff table does not exist")
                active_staff = 0
        except Exception as e:
            current_app.logger.warning(f"Error getting active staff for property {property_id}: {str(e)}")
            active_staff = 0
        
        # Calculate occupancy rate
        occupancy_rate = round((occupied_units / total_units * 100), 2) if total_units > 0 else 0
        
        # Log unit metrics for debugging
        current_app.logger.debug(f"Property {property_id} unit metrics: total={total_units}, occupied={occupied_units}, available={available_units}, rate={occupancy_rate}%")
        
        # Financial metrics - current month
        current_month = datetime.now().month
        current_year = datetime.now().year
        
        # Total income for current month - Property-specific
        try:
            # Check if required tables exist
            if table_exists('payments') and table_exists('bills') and table_exists('units'):
                try:
                    if hasattr(BillStatus, 'PAID') and hasattr(PaymentStatus, 'COMPLETED'):
                        monthly_income = db.session.query(func.sum(Payment.amount)).join(Bill).join(Unit).filter(
                            Unit.property_id == property_id,
                            Bill.status == BillStatus.PAID,
                            extract('month', Payment.payment_date) == current_month,
                            extract('year', Payment.payment_date) == current_year,
                            Payment.status == PaymentStatus.COMPLETED
                        ).scalar() or Decimal('0.00')
                    else:
                        monthly_income = db.session.query(func.sum(Payment.amount)).join(Bill).join(Unit).filter(
                            Unit.property_id == property_id,
                            Bill.status == 'PAID',
                            extract('month', Payment.payment_date) == current_month,
                            extract('year', Payment.payment_date) == current_year,
                            Payment.status == 'COMPLETED'
                        ).scalar() or Decimal('0.00')
                except Exception as join_error:
                    current_app.logger.warning(f"Error joining for monthly income: {str(join_error)}")
                    monthly_income = Decimal('0.00')
            else:
                monthly_income = Decimal('0.00')
        except Exception as e:
            current_app.logger.warning(f"Error getting monthly income: {str(e)}")
            monthly_income = Decimal('0.00')
        
        # Outstanding balance - Property-specific
        # amount_due is a property, so calculate manually
        try:
            if table_exists('bills') and table_exists('units'):
                try:
                    # Get all pending/overdue bills for this property
                    if hasattr(BillStatus, 'PENDING') and hasattr(BillStatus, 'OVERDUE'):
                        bills = Bill.query.join(Unit).filter(
                            Unit.property_id == property_id,
                            Bill.status.in_([BillStatus.PENDING, BillStatus.OVERDUE])
                        ).all()
                    else:
                        bills = Bill.query.join(Unit).filter(
                            Unit.property_id == property_id,
                            Bill.status.in_(['PENDING', 'OVERDUE'])
                        ).all()
                    
                    # Sum amount_due for each bill (amount_due is a calculated property)
                    outstanding_balance = sum(float(bill.amount_due) for bill in bills)
                    outstanding_balance = Decimal(str(outstanding_balance))
                except Exception as join_error:
                    current_app.logger.warning(f"Error joining for outstanding balance: {str(join_error)}")
                    outstanding_balance = Decimal('0.00')
            else:
                current_app.logger.warning("Bills or Units tables do not exist, outstanding balance set to 0.")
                outstanding_balance = Decimal('0.00')
        except Exception as e:
            current_app.logger.warning(f"Error getting outstanding balance: {str(e)}")
            outstanding_balance = Decimal('0.00')
        
        # Recent maintenance requests - Property-specific
        try:
            if table_exists('maintenance_requests') and table_exists('units'):
                try:
                    if hasattr(RequestStatus, 'COMPLETED'):
                        recent_requests = MaintenanceRequest.query.join(Unit).filter(
                            Unit.property_id == property_id,
                            MaintenanceRequest.status != RequestStatus.COMPLETED
                        ).order_by(MaintenanceRequest.created_at.desc()).limit(10).all()
                    else:
                        recent_requests = MaintenanceRequest.query.join(Unit).filter(
                            Unit.property_id == property_id,
                            MaintenanceRequest.status != 'COMPLETED'
                        ).order_by(MaintenanceRequest.created_at.desc()).limit(10).all()
                except Exception as join_error:
                    current_app.logger.warning(f"Error joining maintenance requests: {str(join_error)}")
                    # Fallback: get all maintenance requests
                    recent_requests = MaintenanceRequest.query.limit(10).all()
            else:
                recent_requests = []
        except Exception as e:
            current_app.logger.warning(f"Error getting maintenance requests: {str(e)}")
            recent_requests = []
        
        # Pending tasks - Property-specific (match task filtering logic used in task routes)
        try:
            pending_tasks = []
            if table_exists('tasks'):
                try:
                    unit_ids = []
                    tenant_ids = []
                    
                    if table_exists('units'):
                        unit_ids = [
                            u[0] for u in db.session.query(Unit.id)
                            .filter(Unit.property_id == property_id).all()
                        ]
                    
                    if table_exists('tenants'):
                        tenant_ids = [
                            t[0] for t in db.session.query(Tenant.id)
                            .filter(Tenant.property_id == property_id).all()
                        ]
                    
                    conditions = []
                    if unit_ids:
                        conditions.append(Task.unit_id.in_(unit_ids))
                    if tenant_ids:
                        conditions.append(Task.tenant_id.in_(tenant_ids))
                    
                    query = Task.query.filter(Task.status.in_(['open', 'in_progress']))
                    if conditions:
                        query = query.filter(or_(*conditions) if len(conditions) > 1 else conditions[0])
                    else:
                        query = query.filter(Task.id == -1)
                    
                    pending_tasks = query.order_by(Task.created_at.desc()).limit(10).all()
                except Exception as filter_error:
                    current_app.logger.warning(f"Error filtering pending tasks: {str(filter_error)}")
                    pending_tasks = Task.query.filter(
                        Task.status.in_(['open', 'in_progress'])
                    ).order_by(Task.created_at.desc()).limit(10).all()
        except Exception as e:
            current_app.logger.warning(f"Error getting pending tasks: {str(e)}")
            pending_tasks = []
        
        # Recent announcements - Property-specific (using property_id and is_published)
        try:
            if table_exists('announcements'):
                # Use property_id and is_published (database column names)
                recent_announcements = Announcement.query.filter(
                    or_(
                        Announcement.property_id == property_id,
                        Announcement.property_id.is_(None)  # Include global announcements
                    ),
                    Announcement.is_published == True
                ).order_by(Announcement.created_at.desc()).limit(5).all()
            else:
                recent_announcements = []
        except Exception as e:
            current_app.logger.warning(f"Error getting announcements: {str(e)}")
            recent_announcements = []
        
        # Revenue trend data (last 6 months)
        sales_data = []
        try:
            # Check if required tables exist
            payments_table_exists = table_exists('payments')
            bills_table_exists = table_exists('bills')
            
            if payments_table_exists and bills_table_exists:
                for i in range(5, -1, -1):  # Last 6 months
                    target_date = datetime.now() - timedelta(days=i * 30)
                    month_name = target_date.strftime('%b %Y')
                    
                    try:
                        if hasattr(PaymentStatus, 'COMPLETED'):
                            monthly_revenue = db.session.query(func.sum(Payment.amount)).join(Bill).join(Unit).filter(
                                Unit.property_id == property_id,
                                extract('month', Payment.payment_date) == target_date.month,
                                extract('year', Payment.payment_date) == target_date.year,
                                Payment.status == PaymentStatus.COMPLETED
                            ).scalar() or Decimal('0.00')
                        else:
                            monthly_revenue = db.session.query(func.sum(Payment.amount)).join(Bill).join(Unit).filter(
                                Unit.property_id == property_id,
                                extract('month', Payment.payment_date) == target_date.month,
                                extract('year', Payment.payment_date) == target_date.year,
                                Payment.status == 'COMPLETED'
                            ).scalar() or Decimal('0.00')
                        
                        # Normalize to percentage for chart (assuming max 125k)
                        trend_value = min(float(monthly_revenue) / 1000, 125)  # Convert to thousands
                        actual_value = min(float(monthly_revenue) / 1000, 125)
                        
                        sales_data.append({
                            'month': month_name,
                            'trend': trend_value,
                            'actual': actual_value
                        })
                    except Exception as month_error:
                        current_app.logger.warning(f"Error getting revenue for {month_name}: {str(month_error)}")
                        sales_data.append({
                            'month': month_name,
                            'trend': 0,
                            'actual': 0
                        })
            else:
                # Tables don't exist, return empty data
                current_app.logger.warning("Payments or Bills tables do not exist, returning empty sales data")
                for i in range(5, -1, -1):
                    target_date = datetime.now() - timedelta(days=i * 30)
                    month_name = target_date.strftime('%b %Y')
                    sales_data.append({
                        'month': month_name,
                        'trend': 0,
                        'actual': 0
                    })
        except Exception as e:
            current_app.logger.warning(f"Error generating sales data: {str(e)}")
            # Return empty sales data
            sales_data = []
        
        # Get current month name
        current_month_name = datetime.now().strftime('%B %Y')
        
        # Additional metrics: Bookings today (tenants moved in today)
        bookings_today = 0
        try:
            if TENANT_UNIT_AVAILABLE and TenantUnit and table_exists('tenant_units'):
                today = date.today()
                bookings_today = db.session.query(TenantUnit).join(Unit).filter(
                    Unit.property_id == property_id,
                    func.date(TenantUnit.move_in_date) == today
                ).count()
        except Exception as e:
            current_app.logger.warning(f"Error getting bookings today: {str(e)}")
            bookings_today = 0
        
        # Inquiries this month (from chats with "new inquiry" or "new conversation" subject)
        inquiries_this_month = 0
        try:
            if table_exists('chats'):
                from models.chat import Chat
                current_month_start = datetime(current_year, current_month, 1)
                inquiries_this_month = Chat.query.filter(
                    Chat.property_id == property_id,
                    Chat.created_at >= current_month_start,
                    or_(
                        func.lower(Chat.subject).like('%inquiry%'),
                        func.lower(Chat.subject).like('%conversation%'),
                        Chat.subject.is_(None)
                    )
                ).count()
        except Exception as e:
            current_app.logger.warning(f"Error getting inquiries this month: {str(e)}")
            inquiries_this_month = 0
        
        # Average resolution days (from completed maintenance requests)
        avg_resolution_days = 0
        try:
            if table_exists('maintenance_requests') and table_exists('units'):
                try:
                    if hasattr(RequestStatus, 'COMPLETED'):
                        completed_requests = MaintenanceRequest.query.join(Unit).filter(
                            Unit.property_id == property_id,
                            MaintenanceRequest.status == RequestStatus.COMPLETED,
                            MaintenanceRequest.resolved_at.isnot(None)
                        ).all()
                    else:
                        completed_requests = MaintenanceRequest.query.join(Unit).filter(
                            Unit.property_id == property_id,
                            MaintenanceRequest.status == 'COMPLETED',
                            MaintenanceRequest.resolved_at.isnot(None)
                        ).all()
                    
                    if completed_requests:
                        total_days = 0
                        for req in completed_requests:
                            try:
                                if req.resolved_at and req.created_at:
                                    resolved = req.resolved_at
                                    created = req.created_at
                                    # Ensure both are timezone-aware
                                    if resolved.tzinfo is None:
                                        resolved = resolved.replace(tzinfo=timezone.utc)
                                    if created.tzinfo is None:
                                        created = created.replace(tzinfo=timezone.utc)
                                    days = (resolved - created).days
                                    total_days += days
                            except Exception:
                                continue
                        avg_resolution_days = round(total_days / len(completed_requests), 1) if completed_requests else 0
                except Exception as e:
                    current_app.logger.warning(f"Error calculating avg resolution days: {str(e)}")
                    avg_resolution_days = 0
        except Exception as e:
            current_app.logger.warning(f"Error getting avg resolution days: {str(e)}")
            avg_resolution_days = 0
        
        # Average monthly rent (from bills or units)
        avg_monthly_rent = 0
        try:
            if table_exists('bills') and table_exists('units'):
                # Get average rent from bills for this property
                bills_with_rent = Bill.query.join(Unit).filter(
                    Unit.property_id == property_id
                ).all()
                
                if bills_with_rent:
                    total_rent = sum(float(bill.amount) for bill in bills_with_rent if hasattr(bill, 'amount'))
                    avg_monthly_rent = round(total_rent / len(bills_with_rent), 2) if bills_with_rent else 0
                else:
                    # Fallback: try to get from units if they have rent info
                    try:
                        units_with_rent = Unit.query.filter_by(property_id=property_id).all()
                        if units_with_rent:
                            # If units have monthly_rent field
                            rents = [float(u.monthly_rent) for u in units_with_rent if hasattr(u, 'monthly_rent') and u.monthly_rent]
                            if rents:
                                avg_monthly_rent = round(sum(rents) / len(rents), 2)
                    except Exception:
                        pass
        except Exception as e:
            current_app.logger.warning(f"Error getting avg monthly rent: {str(e)}")
            avg_monthly_rent = 0
        
        # Safely serialize objects to dictionaries
        maintenance_requests_data = []
        for req in recent_requests:
            try:
                maintenance_requests_data.append(req.to_dict(include_tenant=True, include_unit=True))
            except Exception as e:
                current_app.logger.warning(f"Error serializing maintenance request {req.id}: {str(e)}")
                continue
        
        pending_tasks_data = []
        for task in pending_tasks:
            try:
                pending_tasks_data.append(task.to_dict())
            except Exception as e:
                current_app.logger.warning(f"Error serializing task {task.id}: {str(e)}")
                continue
        
        announcements_data = []
        for ann in recent_announcements:
            try:
                announcements_data.append(ann.to_dict())
            except Exception as e:
                current_app.logger.warning(f"Error serializing announcement {ann.id}: {str(e)}")
                continue
        
        # Safely get property name
        property_name = None
        try:
            property_name = getattr(property_obj, 'name', None) or getattr(property_obj, 'title', None) or f'Property {property_id}'
        except Exception:
            property_name = f'Property {property_id}'
        
        # Log final metrics for debugging
        current_app.logger.debug(f"Dashboard data for property {property_id} ({property_name}): "
                              f"revenue={float(monthly_income)}, tenants={active_tenants}, "
                              f"units={total_units}/{occupied_units}/{available_units}, "
                              f"occupancy={occupancy_rate}%")
        
        return jsonify({
            'property_id': property_id,
            'property_name': property_name,
            'metrics': {
                'total_income': float(monthly_income),
                'current_month': current_month_name,
                'active_tenants': active_tenants,
                'active_staff': active_staff,
                'total_properties': total_properties,
                'occupancy_rate': occupancy_rate,
                'outstanding_balance': float(outstanding_balance),
                'bookings_today': bookings_today,
                'inquiries_this_month': inquiries_this_month,
                'avg_resolution_days': avg_resolution_days,
                'avg_monthly_rent': float(avg_monthly_rent)
            },
            'properties': {
                'total': total_properties,
                'total_units': total_units,
                'occupied_units': occupied_units,
                'available_units': available_units,
                'occupancy_rate': occupancy_rate
            },
            'sales_data': sales_data,
            'maintenance_requests': maintenance_requests_data,
            'pending_tasks': pending_tasks_data,
            'announcements': announcements_data
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        current_app.logger.error(f"Manager dashboard error: {str(e)}\n{error_trace}", exc_info=True)
        
        # Return a safe default response instead of crashing
        # This ensures the frontend can still render something
        try:
            property_obj = Property.query.get(property_id) if property_id else None
            property_name = property_obj.name if property_obj else None
        except:
            property_name = None
        
        # Return minimal safe data
        safe_response = {
            'property_id': property_id,
            'property_name': property_name,
            'metrics': {
                'total_income': 0.0,
                'current_month': datetime.now().strftime('%B %Y'),
                'active_tenants': 0,
                'active_staff': 0,
                'total_properties': 1,
                'occupancy_rate': 0.0,
                'outstanding_balance': 0.0
            },
            'properties': {
                'total': 1,
                'total_units': 0,
                'occupied_units': 0,
                'available_units': 0,
                'occupancy_rate': 0.0
            },
            'sales_data': [],
            'maintenance_requests': [],
            'pending_tasks': [],
            'announcements': [],
            'error': 'Some data could not be loaded',
            'error_details': str(e) if current_app.config.get('DEBUG', False) else None
        }
        
        # Add CORS headers manually in case of error
        response = jsonify(safe_response)
        response.status_code = 200  # Return 200 with error message instead of 500
        
        return response

def get_staff_dashboard(user_id, property_id=None):
    """Get staff dashboard data."""
    # Initialize default response early - this will be returned on any error
    default_response = {
        'staff_info': {
            'id': user_id if user_id else None,
            'name': 'Staff Member',
            'email': ''
        },
        'tasks': {
            'pending_count': 0,
            'completed_count': 0,
            'recent_tasks': []
        },
        'maintenance_requests': [],
        'announcements': []
    }
    
    # Wrap everything in try-catch to ensure we always return a response
    try:
        # Convert user_id to int if it's a string
        try:
            user_id_int = int(user_id) if user_id else None
        except (ValueError, TypeError):
            current_app.logger.warning(f"Invalid user_id: {user_id}")
            response = jsonify(default_response)
            response.status_code = 200
            return response
        
        user = None
        try:
            user = User.query.get(user_id_int)
        except Exception as user_query_err:
            current_app.logger.error(f"Error querying user: {str(user_query_err)}", exc_info=True)
            response = jsonify(default_response)
            response.status_code = 200
            return response
        
        if not user:
            current_app.logger.warning(f"User {user_id_int} not found")
            response = jsonify(default_response)
            response.status_code = 200
            return response
        
        # Get staff profile - query directly from Staff table instead of using relationship
        staff = None
        try:
            # Query Staff directly by user_id
            staff = Staff.query.filter_by(user_id=user_id_int).first()
            if not staff and property_id:
                # Also try filtering by property_id if provided
                staff = Staff.query.filter_by(user_id=user_id_int, property_id=property_id).first()
        except Exception as staff_err:
            current_app.logger.warning(f"Error getting staff profile: {str(staff_err)}", exc_info=True)
            staff = None
        
        # If no staff profile, return basic dashboard
        if not staff:
            current_app.logger.debug(f"No staff profile found for user {user_id_int}, returning basic dashboard")
            try:
                user_name = f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip() or getattr(user, 'email', 'Staff Member')
                user_email = getattr(user, 'email', '')
            except Exception:
                user_name = 'Staff Member'
                user_email = ''
            
            response = jsonify({
                'staff_info': {
                    'id': user.id,
                    'name': user_name,
                    'email': user_email
                },
                'tasks': {
                    'pending_count': 0,
                    'completed_count': 0,
                    'recent_tasks': []
                },
                'maintenance_requests': [],
                'announcements': []
            })
            response.status_code = 200
            return response
        
        # My tasks - filter by property if property_id is provided
        # Note: Task.assigned_to references users.id, not staff.id
        # So we need to use staff.user_id (which is the same as user.id)
        try:
            # Get user_id from staff (staff.user_id) or fallback to user.id
            staff_user_id = getattr(staff, 'user_id', None) or user.id
            if not staff_user_id:
                current_app.logger.warning(f"Could not determine user_id for staff {staff.id}")
                staff_user_id = user_id_int
            
            if property_id:
                # Filter tasks by property through unit
                # Tasks can be assigned to users (staff members), so we filter by user_id
                try:
                    my_tasks = Task.query.join(Unit, Task.unit_id == Unit.id).filter(
                        Task.assigned_to == staff_user_id,
                        Unit.property_id == property_id
                    ).order_by(Task.created_at.desc()).limit(10).all()
                except Exception as join_err:
                    # If join fails, try without join (tasks without units)
                    current_app.logger.warning(f"Error joining Unit for tasks: {str(join_err)}")
                    my_tasks = Task.query.filter_by(assigned_to=staff_user_id).order_by(Task.created_at.desc()).limit(10).all()
                
                # Count pending tasks (status is 'open' or 'in_progress', not 'pending')
                try:
                    pending_tasks_count = Task.query.join(Unit, Task.unit_id == Unit.id).filter(
                        Task.assigned_to == staff_user_id,
                        Unit.property_id == property_id,
                        Task.status.in_(['open', 'in_progress'])
                    ).count()
                except Exception:
                    pending_tasks_count = Task.query.filter(
                        Task.assigned_to == staff_user_id,
                        Task.status.in_(['open', 'in_progress'])
                    ).count()
                
                # Count completed tasks
                try:
                    completed_tasks_count = Task.query.join(Unit, Task.unit_id == Unit.id).filter(
                        Task.assigned_to == staff_user_id,
                        Unit.property_id == property_id,
                        Task.status == 'completed'
                    ).count()
                except Exception:
                    completed_tasks_count = Task.query.filter(
                        Task.assigned_to == staff_user_id,
                        Task.status == 'completed'
                    ).count()
            else:
                my_tasks = Task.query.filter_by(assigned_to=staff_user_id).order_by(Task.created_at.desc()).limit(10).all()
                pending_tasks_count = Task.query.filter(
                    Task.assigned_to == staff_user_id,
                    Task.status.in_(['open', 'in_progress'])
                ).count()
                completed_tasks_count = Task.query.filter_by(assigned_to=staff_user_id, status='completed').count()
        except Exception as task_err:
            current_app.logger.warning(f"Error getting tasks: {str(task_err)}", exc_info=True)
            my_tasks = []
            pending_tasks_count = 0
            completed_tasks_count = 0
        
        # My maintenance requests - filter by property if property_id is provided
        # Note: MaintenanceRequest.assigned_to references staff.id (not user.id)
        try:
            if property_id:
                try:
                    my_requests = MaintenanceRequest.query.join(Unit).filter(
                        MaintenanceRequest.assigned_to == staff.id,
                        Unit.property_id == property_id
                    ).order_by(MaintenanceRequest.created_at.desc()).limit(10).all()
                except Exception as join_err:
                    # If join fails, try without join
                    current_app.logger.warning(f"Error joining Unit for requests: {str(join_err)}")
                    my_requests = MaintenanceRequest.query.filter_by(
                        assigned_to=staff.id
                    ).order_by(MaintenanceRequest.created_at.desc()).limit(10).all()
            else:
                my_requests = MaintenanceRequest.query.filter_by(
                    assigned_to=staff.id
                ).order_by(MaintenanceRequest.created_at.desc()).limit(10).all()
        except Exception as req_err:
            current_app.logger.warning(f"Error getting maintenance requests: {str(req_err)}", exc_info=True)
            my_requests = []
        
        # Announcements - filter by property if property_id is provided
        try:
            if property_id:
                recent_announcements = Announcement.query.filter(
                    or_(
                        Announcement.property_id == property_id,
                        Announcement.property_id.is_(None)  # Include global announcements
                    ),
                    Announcement.is_published == True
                ).order_by(Announcement.created_at.desc()).limit(5).all()
            else:
                recent_announcements = Announcement.query.filter_by(
                    is_published=True
                ).order_by(Announcement.created_at.desc()).limit(5).all()
        except Exception as ann_err:
            current_app.logger.warning(f"Error getting announcements: {str(ann_err)}")
            recent_announcements = []
        
        # Safely serialize data
        try:
            staff_info = staff.to_dict(include_user=True) if hasattr(staff, 'to_dict') else {
                'id': staff.id if hasattr(staff, 'id') else None,
                'name': f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip() or user.email,
                'email': user.email
            }
        except Exception:
            staff_info = {
                'id': staff.id if hasattr(staff, 'id') else None,
                'name': f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip() or user.email,
                'email': user.email
            }
        
        tasks_data = []
        for task in my_tasks:
            try:
                tasks_data.append(task.to_dict() if hasattr(task, 'to_dict') else {'id': task.id})
            except Exception:
                continue
        
        requests_data = []
        for req in my_requests:
            try:
                requests_data.append(req.to_dict(include_tenant=True, include_unit=True) if hasattr(req, 'to_dict') else {'id': req.id})
            except Exception:
                continue
        
        announcements_data = []
        for ann in recent_announcements:
            try:
                announcements_data.append(ann.to_dict() if hasattr(ann, 'to_dict') else {'id': ann.id})
            except Exception:
                continue
        
        response_data = {
            'staff_info': staff_info,
            'tasks': {
                'pending_count': pending_tasks_count,
                'completed_count': completed_tasks_count,
                'recent_tasks': tasks_data
            },
            'maintenance_requests': requests_data,
            'announcements': announcements_data
        }
        
        response = jsonify(response_data)
        response.status_code = 200
        return response
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        current_app.logger.error(f"Staff dashboard error: {str(e)}\n{error_trace}", exc_info=True)
        
        # Return safe default response - always return 200 to prevent CORS issues
        safe_response = default_response.copy()
        safe_response['error'] = 'Some data could not be loaded'
        if current_app.config.get('DEBUG', False):
            safe_response['error_details'] = str(e)
        
        response = jsonify(safe_response)
        response.status_code = 200  # Return 200 instead of 500 to prevent frontend crashes and CORS issues
        return response

def get_tenant_dashboard(user_id):
    """Get tenant dashboard data."""
    try:
        user = User.query.get(user_id)
        if not user or not user.tenant_profile:
            return jsonify({'error': 'Tenant profile not found'}), 404
        
        tenant = user.tenant_profile
        
        # Current lease info
        current_rent = tenant.current_rent
        
        # Recent bills
        recent_bills = Bill.query.filter_by(
            tenant_id=tenant.id
        ).order_by(Bill.created_at.desc()).limit(10).all()
        
        # Outstanding balance (amount_due is a property, calculate manually)
        # Use string values since status is now String type
        bills = Bill.query.filter(
            Bill.tenant_id == tenant.id,
            Bill.status.in_(['pending', 'overdue'])
        ).all()
        outstanding_balance = sum(float(bill.amount_due) for bill in bills)
        outstanding_balance = Decimal(str(outstanding_balance))
        
        # My maintenance requests
        my_requests = MaintenanceRequest.query.filter_by(
            tenant_id=tenant.id
        ).order_by(MaintenanceRequest.created_at.desc()).limit(10).all()
        
        # Announcements (using is_published from database)
        recent_announcements = Announcement.query.filter_by(
            is_published=True
        ).order_by(Announcement.created_at.desc()).limit(5).all()
        
        # Payment history
        payment_history = []
        if recent_bills:
            for bill in recent_bills:
                for payment in bill.payments:
                    payment_history.append({
                        'date': payment.payment_date.isoformat(),
                        'amount': float(payment.amount),
                        'bill_title': bill.title,
                        'payment_method': payment.payment_method.value if hasattr(payment.payment_method, 'value') else str(payment.payment_method)
                    })
        
        return jsonify({
            'tenant_info': tenant.to_dict(include_user=True, include_rent=True),
            'current_rent': current_rent.to_dict(include_unit=True) if current_rent else None,
            'financial_summary': {
                'outstanding_balance': float(outstanding_balance),
                'total_paid': tenant.total_rent_paid,
            },
            'recent_bills': [bill.to_dict(include_unit=True) for bill in recent_bills],
            'maintenance_requests': [req.to_dict(include_unit=True) for req in my_requests],
            'announcements': [ann.to_dict() for ann in recent_announcements],
            'payment_history': sorted(payment_history, key=lambda x: x['date'], reverse=True)[:5]
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"Tenant dashboard error: {str(e)}")
        return jsonify({'error': 'Failed to load tenant dashboard'}), 500

@analytics_bp.route('/financial-summary', methods=['GET'])
@jwt_required()
def get_financial_summary():
    """
    Get financial summary
    ---
    tags:
      - Analytics
    summary: Get financial summary for property managers
    description: Get financial summary for property managers - property-specific
    security:
      - Bearer: []
    parameters:
      - in: query
        name: start_date
        type: string
        format: date
      - in: query
        name: end_date
        type: string
        format: date
    responses:
      200:
        description: Financial summary retrieved successfully
        schema:
          type: object
          properties:
            total_revenue:
              type: number
            total_expenses:
              type: number
            net_income:
              type: number
            by_category:
              type: object
      401:
        description: Unauthorized
      403:
        description: Forbidden - Property Manager access required
      500:
        description: Server error
    """
    try:
        # Get property_id from request (subdomain, query param, header, or JWT)
        property_id = get_property_id_from_request()
        current_app.logger.debug(f"Financial summary request - property_id: {property_id}")
        
        if not property_id:
            current_app.logger.warning("Financial summary: No property_id found in request")
            return jsonify({
                'error': 'Property ID is required. Please access through your property subdomain.'
            }), 400
        
        # Verify property exists - use raw query to avoid enum validation issues
        try:
            # Use raw SQL to get property without triggering enum validation
            property_result = db.session.execute(text(
                "SELECT id, title, description FROM properties WHERE id = :property_id LIMIT 1"
            ), {'property_id': property_id}).first()
            
            if not property_result:
                return jsonify({'error': 'Property not found'}), 404
            
            # Try to get Property object, but if enum validation fails, use raw result
            try:
                property_obj = Property.query.get(property_id)
            except Exception as enum_error:
                # If enum validation fails, create a mock object from raw query
                current_app.logger.warning(f"Enum validation error for property {property_id}, using raw query result: {str(enum_error)}")
                property_obj = type('Property', (), {
                    'id': property_result[0],
                    'name': property_result[1] if property_result[1] else f'Property {property_result[0]}',
                    'title': property_result[1] if property_result[1] else f'Property {property_result[0]}'
                })()
        except Exception as prop_error:
            current_app.logger.error(f"Error getting property {property_id}: {str(prop_error)}", exc_info=True)
            return jsonify({'error': 'Property not found'}), 404
        
        # Monthly revenue for the last 12 months - Property-specific
        monthly_data = []
        for i in range(11, -1, -1):
            target_date = datetime.now() - timedelta(days=i * 30)
            month_name = target_date.strftime('%b %Y')
            
            try:
                # Check if required tables exist
                if table_exists('payments') and table_exists('bills') and table_exists('units'):
                    # Filter by property_id through Unit join
                    monthly_revenue = db.session.query(func.sum(Payment.amount)).join(Bill, Bill.id == Payment.bill_id).join(Unit, Unit.id == Bill.unit_id).filter(
                        Unit.property_id == property_id,
                        extract('month', Payment.payment_date) == target_date.month,
                        extract('year', Payment.payment_date) == target_date.year,
                        Payment.status == 'completed'  # Use lowercase string value
                    ).scalar() or Decimal('0.00')
                else:
                    monthly_revenue = Decimal('0.00')
            except Exception as e:
                current_app.logger.warning(f"Error getting monthly revenue for {month_name}: {str(e)}", exc_info=True)
                monthly_revenue = Decimal('0.00')
            
            monthly_data.append({
                'month': month_name,
                'revenue': float(monthly_revenue)
            })
        
        # Total metrics - Property-specific
        try:
            if table_exists('payments') and table_exists('bills') and table_exists('units'):
                total_revenue = db.session.query(func.sum(Payment.amount)).join(Bill, Bill.id == Payment.bill_id).join(Unit, Unit.id == Bill.unit_id).filter(
                    Unit.property_id == property_id,
                    Payment.status == 'completed'  # Use lowercase string value
                ).scalar() or Decimal('0.00')
            else:
                total_revenue = Decimal('0.00')
        except Exception as e:
            current_app.logger.warning(f"Error getting total revenue: {str(e)}", exc_info=True)
            total_revenue = Decimal('0.00')
        
        # Outstanding balance - Property-specific
        try:
            if table_exists('bills') and table_exists('units'):
                bills = Bill.query.join(Unit, Unit.id == Bill.unit_id).filter(
                    Unit.property_id == property_id,
                    Bill.status.in_(['pending', 'overdue'])
                ).all()
                # Safely calculate outstanding balance
                total_outstanding = 0.0
                for bill in bills:
                    try:
                        # amount_due is a property that queries the database, so handle carefully
                        # Calculate directly: amount - amount_paid
                        bill_amount = float(bill.amount) if hasattr(bill, 'amount') else 0.0
                        
                        # Calculate amount_paid from payments directly (more efficient)
                        try:
                            from models.bill import Payment
                            from sqlalchemy import func
                            amount_paid = db.session.query(func.sum(Payment.amount)).filter(
                                Payment.bill_id == bill.id,
                                Payment.status.in_(['completed', 'approved'])
                            ).scalar() or 0.0
                            amount_paid = float(amount_paid)
                        except Exception:
                            amount_paid = 0.0
                        
                        amount_due = max(0.0, bill_amount - amount_paid)
                        total_outstanding += amount_due
                    except Exception as bill_error:
                        current_app.logger.warning(f"Error calculating amount_due for bill {bill.id}: {str(bill_error)}")
                        continue
                total_outstanding = Decimal(str(total_outstanding))
            else:
                total_outstanding = Decimal('0.00')
        except Exception as e:
            current_app.logger.warning(f"Error getting outstanding balance: {str(e)}", exc_info=True)
            total_outstanding = Decimal('0.00')
        
        # Overdue bills - Property-specific
        try:
            if table_exists('bills') and table_exists('units'):
                overdue_bills = Bill.query.join(Unit, Unit.id == Bill.unit_id).filter(
                    Unit.property_id == property_id,
                    Bill.status == 'overdue'
                ).count()
            else:
                overdue_bills = 0
        except Exception as e:
            current_app.logger.warning(f"Error getting overdue bills count: {str(e)}", exc_info=True)
            overdue_bills = 0
        
        # Safely get property name
        property_name = None
        try:
            property_name = getattr(property_obj, 'name', None) or getattr(property_obj, 'title', None) or f'Property {property_id}'
        except Exception:
            property_name = f'Property {property_id}'
        
        # Log financial summary for debugging
        current_app.logger.debug(f"Financial summary for property {property_id} ({property_name}): "
                              f"total_revenue={float(total_revenue)}, outstanding={float(total_outstanding)}, "
                              f"overdue_bills={overdue_bills}, months={len(monthly_data)}")
        
        return jsonify({
            'property_id': property_id,
            'property_name': property_name,
            'monthly_revenue': monthly_data,
            'totals': {
                'total_revenue': float(total_revenue),
                'outstanding_balance': float(total_outstanding),
                'overdue_bills_count': overdue_bills
            }
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        current_app.logger.error(f"Financial summary error: {str(e)}\n{error_trace}", exc_info=True)
        
        # Return detailed error in DEBUG mode
        error_response = {'error': 'Failed to load financial summary'}
        if current_app.config.get('DEBUG', False):
            error_response['details'] = str(e)
            error_response['traceback'] = error_trace.split('\n')[-5:]  # Last 5 lines
        
        return jsonify(error_response), 500

@analytics_bp.route('/occupancy-report', methods=['GET'])
@jwt_required()
def get_occupancy_report():
    """
    Get occupancy report
    ---
    tags:
      - Analytics
    summary: Get occupancy report for property managers
    description: Get occupancy report for property managers - property-specific
    security:
      - Bearer: []
    parameters:
      - in: query
        name: start_date
        type: string
        format: date
      - in: query
        name: end_date
        type: string
        format: date
    responses:
      200:
        description: Occupancy report retrieved successfully
        schema:
          type: object
          properties:
            total_units:
              type: integer
            occupied_units:
              type: integer
            vacant_units:
              type: integer
            occupancy_rate:
              type: number
            by_unit_type:
              type: object
      401:
        description: Unauthorized
      403:
        description: Forbidden - Property Manager access required
      500:
        description: Server error
    """
    try:
        # Get property_id from request (subdomain, query param, header, or JWT)
        property_id = get_property_id_from_request()
        current_app.logger.debug(f"Occupancy report request - property_id: {property_id}")
        
        if not property_id:
            current_app.logger.warning("Occupancy report: No property_id found in request")
            return jsonify({
                'error': 'Property ID is required. Please access through your property subdomain.'
            }), 400
        
        # Verify property exists - use raw query to avoid enum validation issues
        try:
            # Use raw SQL to get property without triggering enum validation
            property_result = db.session.execute(text(
                "SELECT id, title, description FROM properties WHERE id = :property_id LIMIT 1"
            ), {'property_id': property_id}).first()
            
            if not property_result:
                return jsonify({'error': 'Property not found'}), 404
            
            property_obj = Property.query.get(property_id) if property_result else None
            # If Property.query.get fails due to enum validation, use the raw result
            if not property_obj:
                property_obj = type('Property', (), {
                    'id': property_result[0],
                    'name': property_result[1] if property_result[1] else f'Property {property_result[0]}',
                    'title': property_result[1] if property_result[1] else f'Property {property_result[0]}'
                })()
        except Exception as prop_error:
            current_app.logger.warning(f"Error getting property: {str(prop_error)}")
            # Fallback: use raw query result
            try:
                property_result = db.session.execute(text(
                    "SELECT id, title FROM properties WHERE id = :property_id LIMIT 1"
                ), {'property_id': property_id}).first()
                if property_result:
                    property_obj = type('Property', (), {
                        'id': property_result[0],
                        'name': property_result[1] if property_result[1] else f'Property {property_result[0]}',
                        'title': property_result[1] if property_result[1] else f'Property {property_result[0]}'
                    })()
                else:
                    return jsonify({'error': 'Property not found'}), 404
            except Exception:
                return jsonify({'error': 'Property not found'}), 404
        
        # Overall occupancy - Property-specific
        # Calculate based on actual tenant-unit relationships for accuracy
        try:
            if table_exists('units'):
                total_units = Unit.query.filter_by(property_id=property_id).count()
                
                # Calculate occupied units based on TenantUnit relationships
                occupied_units = 0
                if TENANT_UNIT_AVAILABLE and TenantUnit and table_exists('tenant_units'):
                    try:
                        today = date.today()
                        
                        # Get distinct units that have active tenants (occupied units)
                        occupied_unit_ids = db.session.query(Unit.id).join(TenantUnit).filter(
                            Unit.property_id == property_id,
                            TenantUnit.move_in_date.isnot(None),
                            or_(
                                TenantUnit.move_out_date.is_(None),  # Ongoing lease (no move-out date)
                                TenantUnit.move_out_date >= today   # Future move-out date
                            )
                        ).distinct().all()
                        
                        occupied_units = len(occupied_unit_ids)
                        current_app.logger.debug(f"Occupancy report: {occupied_units} occupied units (from TenantUnit) for property {property_id}")
                    except Exception as join_error:
                        current_app.logger.warning(f"Error joining TenantUnit in occupancy report: {str(join_error)}")
                        # Fallback to unit status
                        occupied_units = Unit.query.filter(
                            Unit.property_id == property_id,
                            or_(Unit.status == 'occupied', Unit.status == 'rented')
                        ).count()
                else:
                    # Fallback to unit status if TenantUnit not available
                    occupied_units = Unit.query.filter(
                        Unit.property_id == property_id,
                        or_(Unit.status == 'occupied', Unit.status == 'rented')
                    ).count()
                
                # Available units = total - occupied
                available_units = max(0, total_units - occupied_units)
                occupancy_rate = round((occupied_units / total_units * 100), 2) if total_units > 0 else 0
            else:
                total_units = 0
                occupied_units = 0
                available_units = 0
                occupancy_rate = 0
        except Exception as e:
            current_app.logger.warning(f"Error getting occupancy data: {str(e)}", exc_info=True)
            total_units = 0
            occupied_units = 0
            available_units = 0
            occupancy_rate = 0
        
        # Unit type breakdown (if property_type info available)
        unit_type_data = []
        try:
            # Group by property_type or similar if available
            # For now, we'll use a simple breakdown
            unit_type_data = [
                {
                    'type': 'All Units',
                    'total': total_units,
                    'occupied': occupied_units,
                    'available': available_units,
                    'occupancy_rate': occupancy_rate
                }
            ]
        except Exception as e:
            current_app.logger.warning(f"Error getting unit type breakdown: {str(e)}")
            unit_type_data = []
        
        # Safely get property name
        property_name = None
        try:
            property_name = getattr(property_obj, 'name', None) or getattr(property_obj, 'title', None) or f'Property {property_id}'
        except Exception:
            property_name = f'Property {property_id}'
        
        # Log occupancy report for debugging
        current_app.logger.debug(f"Occupancy report for property {property_id} ({property_name}): "
                              f"total={total_units}, occupied={occupied_units}, "
                              f"available={available_units}, rate={occupancy_rate}%, "
                              f"breakdown_items={len(unit_type_data)}")
        
        return jsonify({
            'property_id': property_id,
            'property_name': property_name,
            'overall_occupancy': {
                'total_units': total_units,
                'occupied_units': occupied_units,
                'available_units': available_units,
                'occupancy_rate': occupancy_rate
            },
            'unit_type_breakdown': unit_type_data
        }), 200
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        current_app.logger.error(f"Occupancy report error: {str(e)}\n{error_trace}", exc_info=True)
        
        # Return detailed error in DEBUG mode
        error_response = {'error': 'Failed to load occupancy report'}
        if current_app.config.get('DEBUG', False):
            error_response['details'] = str(e)
            error_response['traceback'] = error_trace.split('\n')[-5:]  # Last 5 lines
        
        return jsonify(error_response), 500


def _get_analytics_data_for_report(property_id):
    """
    Helper function to get analytics data for reports.
    Directly queries database to avoid Flask request context issues.
    Matches frontend data extraction logic exactly.
    """
    try:
        current_app.logger.debug(f'Getting analytics data for report, property_id: {property_id}')
        
        # Get property name
        property_name = f'Property {property_id}'
        try:
            property_result = db.session.execute(text(
                "SELECT id, title, name FROM properties WHERE id = :property_id LIMIT 1"
            ), {'property_id': property_id}).first()
            if property_result:
                property_name = property_result[1] or property_result[2] if len(property_result) > 2 else property_result[1] or f'Property {property_id}'
        except Exception as e:
            current_app.logger.warning(f'Error getting property name: {str(e)}')
        
        # Get dashboard data - this should return the same data the frontend sees
        try:
            dashboard_data = get_manager_dashboard(property_id)
            # Check if it's a response object
            if hasattr(dashboard_data, 'get_json'):
                dashboard_json = dashboard_data.get_json()
                # Check status code
                if hasattr(dashboard_data, 'status_code') and dashboard_data.status_code != 200:
                    current_app.logger.error(f'Dashboard returned status {dashboard_data.status_code}: {dashboard_json}')
                    # Don't set to empty, try to use what we got
            elif isinstance(dashboard_data, dict):
                dashboard_json = dashboard_data
            else:
                current_app.logger.error(f'Dashboard returned unexpected type: {type(dashboard_data)}')
                dashboard_json = {}
        except Exception as e:
            current_app.logger.error(f'Error getting dashboard data: {str(e)}', exc_info=True)
            dashboard_json = {}
        
        # Log the FULL dashboard response to debug
        current_app.logger.debug(f'Dashboard JSON keys: {list(dashboard_json.keys()) if dashboard_json else "EMPTY"}')
        if dashboard_json:
            current_app.logger.debug(f'Dashboard has error: {dashboard_json.get("error")}')
            if dashboard_json.get('metrics'):
                current_app.logger.debug(f'Dashboard metrics keys: {list(dashboard_json.get("metrics", {}).keys())}')
                current_app.logger.debug(f'Dashboard metrics sample: total_income={dashboard_json.get("metrics", {}).get("total_income")}, avg_monthly_rent={dashboard_json.get("metrics", {}).get("avg_monthly_rent")}')
            if dashboard_json.get('properties'):
                current_app.logger.debug(f'Dashboard properties keys: {list(dashboard_json.get("properties", {}).keys())}')
                current_app.logger.debug(f'Dashboard properties sample: total_units={dashboard_json.get("properties", {}).get("total_units")}, occupied_units={dashboard_json.get("properties", {}).get("occupied_units")}')
        
        # If dashboard has error, log it but continue
        if dashboard_json.get('error'):
            current_app.logger.error(f'Dashboard returned error: {dashboard_json.get("error")}')
        
        # Extract dashboard metrics and properties - be very explicit
        metrics = {}
        properties = {}
        if dashboard_json:
            metrics = dashboard_json.get('metrics') or {}
            properties = dashboard_json.get('properties') or {}
        
        # Log what we extracted
        current_app.logger.debug(f'Extracted metrics keys: {list(metrics.keys())}, properties keys: {list(properties.keys())}')
        if metrics:
            current_app.logger.debug(f'Metrics values: total_income={metrics.get("total_income")}, active_tenants={metrics.get("active_tenants")}, avg_monthly_rent={metrics.get("avg_monthly_rent")}, occupancy_rate={metrics.get("occupancy_rate")}, outstanding_balance={metrics.get("outstanding_balance")}')
        if properties:
            current_app.logger.debug(f'Properties values: total_units={properties.get("total_units")}, occupied_units={properties.get("occupied_units")}, available_units={properties.get("available_units")}, occupancy_rate={properties.get("occupancy_rate")}')
        
        # Get financial data directly from database (same logic as get_financial_summary)
        financial_totals = {}
        try:
            # Total revenue - all completed payments
            if table_exists('payments') and table_exists('bills') and table_exists('units'):
                total_revenue = db.session.query(func.sum(Payment.amount)).join(Bill, Bill.id == Payment.bill_id).join(Unit, Unit.id == Bill.unit_id).filter(
                    Unit.property_id == property_id,
                    Payment.status == 'completed'
                ).scalar() or Decimal('0.00')
            else:
                total_revenue = Decimal('0.00')
            
            # Outstanding balance
            total_outstanding = Decimal('0.00')
            if table_exists('bills') and table_exists('units'):
                bills = Bill.query.join(Unit, Unit.id == Bill.unit_id).filter(
                    Unit.property_id == property_id,
                    Bill.status.in_(['pending', 'overdue'])
                ).all()
                for bill in bills:
                    try:
                        bill_amount = float(bill.amount) if hasattr(bill, 'amount') else 0.0
                        amount_paid = db.session.query(func.sum(Payment.amount)).filter(
                            Payment.bill_id == bill.id,
                            Payment.status.in_(['completed', 'approved'])
                        ).scalar() or 0.0
                        amount_due = max(0.0, bill_amount - float(amount_paid))
                        total_outstanding += Decimal(str(amount_due))
                    except Exception:
                        continue
            
            # Overdue bills count
            overdue_bills = 0
            if table_exists('bills') and table_exists('units'):
                overdue_bills = Bill.query.join(Unit, Unit.id == Bill.unit_id).filter(
                    Unit.property_id == property_id,
                    Bill.status == 'overdue'
                ).count()
            
            financial_totals = {
                'total_revenue': float(total_revenue),
                'outstanding_balance': float(total_outstanding),
                'overdue_bills_count': overdue_bills
            }
            current_app.logger.debug(f'Financial totals: {financial_totals}')
        except Exception as e:
            current_app.logger.warning(f'Error getting financial data: {str(e)}', exc_info=True)
        
        # Get occupancy data directly from database (same logic as get_occupancy_report)
        overall_occupancy = {}
        try:
            if table_exists('units'):
                total_units = Unit.query.filter_by(property_id=property_id).count()
                
                # Calculate occupied units
                occupied_units = 0
                if TENANT_UNIT_AVAILABLE and TenantUnit and table_exists('tenant_units'):
                    try:
                        today = date.today()
                        occupied_unit_ids = db.session.query(Unit.id).join(TenantUnit).filter(
                            Unit.property_id == property_id,
                            TenantUnit.move_in_date.isnot(None),
                            or_(
                                TenantUnit.move_out_date.is_(None),
                                TenantUnit.move_out_date >= today
                            )
                        ).distinct().all()
                        occupied_units = len(occupied_unit_ids)
                    except Exception:
                        occupied_units = Unit.query.filter(
                            Unit.property_id == property_id,
                            or_(Unit.status == 'occupied', Unit.status == 'rented')
                        ).count()
                else:
                    occupied_units = Unit.query.filter(
                        Unit.property_id == property_id,
                        or_(Unit.status == 'occupied', Unit.status == 'rented')
                    ).count()
                
                available_units = max(0, total_units - occupied_units)
                occupancy_rate = round((occupied_units / total_units * 100), 2) if total_units > 0 else 0
                
                overall_occupancy = {
                    'total_units': total_units,
                    'occupied_units': occupied_units,
                    'available_units': available_units,
                    'occupancy_rate': occupancy_rate
                }
                current_app.logger.debug(f'Occupancy data: {overall_occupancy}')
            else:
                overall_occupancy = {
                    'total_units': 0,
                    'occupied_units': 0,
                    'available_units': 0,
                    'occupancy_rate': 0
                }
        except Exception as e:
            current_app.logger.warning(f'Error getting occupancy data: {str(e)}', exc_info=True)
            overall_occupancy = {
                'total_units': 0,
                'occupied_units': 0,
                'available_units': 0,
                'occupancy_rate': 0
            }
        
        # Use property name from dashboard if available
        if dashboard_json.get('property_name'):
            property_name = dashboard_json.get('property_name')
        
        # Log extracted data from all sources
        current_app.logger.debug(f'All data sources - metrics keys: {list(metrics.keys())}, properties keys: {list(properties.keys())}, financial_totals keys: {list(financial_totals.keys())}, overall_occupancy keys: {list(overall_occupancy.keys())}')
        
        # Build report data with proper defaults - prioritize dashboard data
        # Dashboard has the most comprehensive data, use it as primary source
        # Use explicit None checks to preserve 0 values (0 is valid data, None is not)
        # Match frontend logic: dashboardMetrics?.total_income ?? financialTotals?.total_revenue ?? 0
        def safe_get(d, key, default=None):
            """Safely get value, preserving 0 but defaulting None to default"""
            if not d:
                return default
            val = d.get(key)
            return val if val is not None else default
        
        # Use nullish coalescing pattern (Python equivalent of ??)
        # If first value is None, use second, if that's None, use third
        def coalesce(*values):
            """Return first non-None value, or last value if all are None"""
            for val in values:
                if val is not None:
                    return val
            return values[-1] if values else None
        
        # Extract values - MATCH FRONTEND LOGIC EXACTLY
        # Frontend uses: dashboardMetrics?.total_income ?? financialTotals?.total_revenue ?? 0
        # So we prioritize dashboardMetrics.total_income (current month) first, then financial summary
        total_revenue_from_dashboard = metrics.get('total_income') if metrics else None
        total_revenue_from_financial = financial_totals.get('total_revenue') if financial_totals else None
        total_revenue_val = coalesce(
            total_revenue_from_dashboard,  # First priority: current month from dashboard (matches frontend)
            total_revenue_from_financial,  # Second priority: all time from financial summary
            0
        )
        
        # Log which source we're using for total revenue
        if total_revenue_from_dashboard is not None:
            current_app.logger.debug(f'Using total_income from dashboard (current month): {total_revenue_from_dashboard}')
        elif total_revenue_from_financial is not None:
            current_app.logger.debug(f'Using total_revenue from financial summary (all time): {total_revenue_from_financial}')
        else:
            current_app.logger.warning('No total revenue data found, using 0')
        # MATCH FRONTEND LOGIC EXACTLY - same priority order
        # Frontend: dashboardMetrics?.outstanding_balance ?? financialTotals?.outstanding_balance ?? 0
        outstanding_balance_val = coalesce(
            metrics.get('outstanding_balance') if metrics else None,
            financial_totals.get('outstanding_balance') if financial_totals else None,
            0
        )
        # Frontend: financialTotals?.overdue_bills_count ?? 0
        overdue_bills_val = coalesce(
            financial_totals.get('overdue_bills_count') if financial_totals else None,
            0
        )
        # Frontend: overallOccupancy?.occupancy_rate ?? dashboardProperties?.occupancy_rate ?? dashboardMetrics?.occupancy_rate
        # Then calculates if null and totalUnits > 0: (occupiedUnits / totalUnits) * 100
        occupancy_rate_val = coalesce(
            overall_occupancy.get('occupancy_rate') if overall_occupancy else None,
            properties.get('occupancy_rate') if properties else None,
            metrics.get('occupancy_rate') if metrics else None,
            None  # Don't default to 0 yet, we'll calculate if needed
        )
        # Frontend: overallOccupancy?.total_units ?? dashboardProperties?.total_units ?? 0
        total_units_val = coalesce(
            overall_occupancy.get('total_units') if overall_occupancy else None,
            properties.get('total_units') if properties else None,
            0
        )
        # Frontend: overallOccupancy?.occupied_units ?? dashboardProperties?.occupied_units ?? 0
        occupied_units_val = coalesce(
            overall_occupancy.get('occupied_units') if overall_occupancy else None,
            properties.get('occupied_units') if properties else None,
            0
        )
        # Frontend: overallOccupancy?.available_units ?? dashboardProperties?.available_units ?? 0
        available_units_val = coalesce(
            overall_occupancy.get('available_units') if overall_occupancy else None,
            properties.get('available_units') if properties else None,
            0
        )
        # Calculate occupancy rate if not provided (matches frontend logic)
        if occupancy_rate_val is None:
            if total_units_val and total_units_val > 0:
                occupancy_rate_val = (occupied_units_val / total_units_val) * 100
            else:
                occupancy_rate_val = 0
        
        # Frontend: dashboardMetrics?.active_tenants (no fallback, defaults to 0 in getNumericValue)
        active_tenants_val = coalesce(
            metrics.get('active_tenants') if metrics else None,
            0
        )
        # Frontend: dashboard?.maintenance_requests ? dashboard.maintenance_requests.length : 0
        # But we need TOTAL count, not just recent ones. Query directly for accurate count.
        open_requests_val = 0
        try:
            if table_exists('maintenance_requests') and table_exists('units'):
                if hasattr(RequestStatus, 'COMPLETED'):
                    open_requests_val = MaintenanceRequest.query.join(Unit).filter(
                        Unit.property_id == property_id,
                        MaintenanceRequest.status != RequestStatus.COMPLETED
                    ).count()
                else:
                    open_requests_val = MaintenanceRequest.query.join(Unit).filter(
                        Unit.property_id == property_id,
                        MaintenanceRequest.status != 'COMPLETED'
                    ).count()
            else:
                # Fallback to dashboard data if query fails
                open_requests_val = len(dashboard_json.get('maintenance_requests', []) or [])
        except Exception as e:
            current_app.logger.warning(f'Error getting open requests count: {str(e)}')
            # Fallback to dashboard data
            open_requests_val = len(dashboard_json.get('maintenance_requests', []) or [])
        
        # Frontend: dashboard?.pending_tasks ? dashboard.pending_tasks.length : 0
        pending_tasks_val = len(dashboard_json.get('pending_tasks', []) or [])
        
        # Frontend: dashboardMetrics?.avg_monthly_rent (no fallback shown, but we'll default to 0)
        # Get from metrics first, but also try to calculate if not available
        avg_monthly_rent_val = coalesce(
            metrics.get('avg_monthly_rent') if metrics else None,
            None  # Don't default yet, try to calculate
        )
        
        # If avg_monthly_rent not in metrics, calculate it directly
        if avg_monthly_rent_val is None:
            try:
                if table_exists('bills') and table_exists('units'):
                    bills_with_rent = Bill.query.join(Unit).filter(
                        Unit.property_id == property_id
                    ).all()
                    if bills_with_rent:
                        total_rent = sum(float(bill.amount) for bill in bills_with_rent if hasattr(bill, 'amount'))
                        avg_monthly_rent_val = round(total_rent / len(bills_with_rent), 2) if bills_with_rent else 0
                    else:
                        # Fallback: try to get from units
                        units_with_rent = Unit.query.filter_by(property_id=property_id).all()
                        if units_with_rent:
                            rents = [float(u.monthly_rent) for u in units_with_rent if hasattr(u, 'monthly_rent') and u.monthly_rent]
                            if rents:
                                avg_monthly_rent_val = round(sum(rents) / len(rents), 2)
                            else:
                                avg_monthly_rent_val = 0
                        else:
                            avg_monthly_rent_val = 0
                else:
                    avg_monthly_rent_val = 0
            except Exception as e:
                current_app.logger.warning(f'Error calculating avg monthly rent: {str(e)}')
                avg_monthly_rent_val = 0
        
        current_app.logger.debug(f'Open requests: {open_requests_val}, Avg monthly rent: {avg_monthly_rent_val}')
        
        # Log raw values before processing
        current_app.logger.debug(f'Raw dashboard values: total_income={metrics.get("total_income")}, active_tenants={metrics.get("active_tenants")}, total_units={properties.get("total_units")}, occupied_units={properties.get("occupied_units")}')
        current_app.logger.debug(f'Financial totals: total_revenue={financial_totals.get("total_revenue")}, overdue_bills={financial_totals.get("overdue_bills_count")}')
        current_app.logger.debug(f'Overall occupancy: total_units={overall_occupancy.get("total_units")}, occupied_units={overall_occupancy.get("occupied_units")}, occupancy_rate={overall_occupancy.get("occupancy_rate")}')
        current_app.logger.debug(f'Processed values: total_revenue={total_revenue_val}, active_tenants={active_tenants_val}, total_units={total_units_val}, occupied_units={occupied_units_val}')
        
        # Convert to proper types - preserve 0 values, only default None
        def to_float(val, default=0.0):
            if val is None:
                return default
            try:
                return float(val)
            except (ValueError, TypeError):
                return default
        
        def to_int(val, default=0):
            if val is None:
                return default
            try:
                return int(val)
            except (ValueError, TypeError):
                return default
        
        report_data = {
            'property_id': property_id,
            'property_name': property_name or 'Property Analytics',
            'total_revenue': to_float(total_revenue_val, 0.0),
            'outstanding_balance': to_float(outstanding_balance_val, 0.0),
            'overdue_bills': to_int(overdue_bills_val, 0),
            'occupancy_rate': to_float(occupancy_rate_val, 0.0),
            'total_units': to_int(total_units_val, 0),
            'occupied_units': to_int(occupied_units_val, 0),
            'available_units': to_int(available_units_val, 0),
            'open_requests': to_int(open_requests_val, 0),
            'pending_tasks': to_int(pending_tasks_val, 0),
            'avg_monthly_rent': to_float(avg_monthly_rent_val, 0.0),
            'monthly_revenue': [],  # Not needed for basic report, can be added later if needed
            'unit_type_breakdown': [],  # Not needed for basic report, can be added later if needed
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        # Log final report data summary with ALL values
        current_app.logger.debug(f'FINAL Report data: property_name={report_data["property_name"]}, revenue={report_data["total_revenue"]}, outstanding={report_data["outstanding_balance"]}, overdue_bills={report_data["overdue_bills"]}')
        current_app.logger.debug(f'FINAL Report data: units={report_data["total_units"]}, occupied={report_data["occupied_units"]}, available={report_data["available_units"]}, occupancy={report_data["occupancy_rate"]}%')
        current_app.logger.debug(f'FINAL Report data: open_requests={report_data["open_requests"]}, pending_tasks={report_data["pending_tasks"]}, avg_rent={report_data["avg_monthly_rent"]}')
        
        return report_data
    except Exception as e:
        current_app.logger.error(f'Error getting analytics data for report: {e}', exc_info=True)
        return {
            'property_id': property_id,
            'property_name': 'Property Analytics',
            'total_revenue': 0.0,
            'outstanding_balance': 0.0,
            'overdue_bills': 0,
            'occupancy_rate': 0.0,
            'total_units': 0,
            'occupied_units': 0,
            'available_units': 0,
            'active_tenants': 0,
            'open_requests': 0,
            'pending_tasks': 0,
            'inquiries_this_month': 0,
            'avg_monthly_rent': 0.0,
            'monthly_revenue': [],
            'unit_type_breakdown': [],
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }


@analytics_bp.route('/download/pdf', methods=['GET'])
@jwt_required()
def download_pdf_report():
    """
    Download PDF report
    ---
    tags:
      - Analytics
    summary: Download analytics report as PDF
    description: Download analytics report as PDF. Property Manager only.
    security:
      - Bearer: []
    parameters:
      - in: query
        name: report_type
        type: string
        enum: [dashboard, financial, occupancy]
        default: dashboard
      - in: query
        name: start_date
        type: string
        format: date
      - in: query
        name: end_date
        type: string
        format: date
    responses:
      200:
        description: PDF report file
        schema:
          type: file
      401:
        description: Unauthorized
      403:
        description: Forbidden - Property Manager access required
      500:
        description: Server error
    """
    if not REPORTLAB_AVAILABLE:
        return jsonify({'error': 'PDF generation not available. Please install reportlab.'}), 503
    
    try:
        property_id = get_property_id_from_request()
        if not property_id:
            current_app.logger.error('PDF download: Property ID not found in request')
            return jsonify({'error': 'Property ID is required'}), 400
        
        current_app.logger.debug(f'PDF download: Starting report generation for property_id={property_id}')
        
        data = _get_analytics_data_for_report(property_id)
        
        # Log ALL data for debugging
        current_app.logger.debug(f'PDF Report Data - Full: {data}')
        current_app.logger.debug(f'PDF Report Data - Summary: property_id={property_id}, property_name={data.get("property_name")}, total_revenue={data.get("total_revenue")}, total_units={data.get("total_units")}, active_tenants={data.get("active_tenants")}, occupancy_rate={data.get("occupancy_rate")}')
        
        # Validate data
        if not data:
            current_app.logger.error('No data returned from _get_analytics_data_for_report')
            return jsonify({'error': 'No data available for report generation'}), 400
        
        # Ensure we have at least some data - validate all critical fields
        if data.get('property_name') is None:
            current_app.logger.warning('Property name is None, using default')
            data['property_name'] = 'Property Analytics'
        
        # Ensure all numeric fields are valid numbers (not None)
        numeric_fields = ['total_revenue', 'outstanding_balance', 'occupancy_rate', 'total_units', 'occupied_units', 'available_units', 'overdue_bills', 'open_requests', 'pending_tasks', 'avg_monthly_rent']
        for field in numeric_fields:
            if data.get(field) is None:
                current_app.logger.warning(f'Field {field} is None, setting to 0')
                data[field] = 0
        
        # Log validated data
        current_app.logger.debug(f'PDF Report Data - After validation: revenue={data.get("total_revenue")}, units={data.get("total_units")}, tenants={data.get("active_tenants")}')
        
        # Create PDF in memory
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Property Analytics Report", title_style))
        
        # Report info
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#666666'),
            alignment=1
        )
        story.append(Paragraph(f"Property: {data['property_name']}", info_style))
        story.append(Paragraph(f"Generated: {data['generated_at']}", info_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Key Metrics Table - ensure all values are properly formatted
        # Safely get values with defaults
        total_revenue = float(data.get('total_revenue', 0) or 0)
        outstanding_balance = float(data.get('outstanding_balance', 0) or 0)
        overdue_bills = int(data.get('overdue_bills', 0) or 0)
        occupancy_rate = float(data.get('occupancy_rate', 0) or 0)
        total_units = int(data.get('total_units', 0) or 0)
        occupied_units = int(data.get('occupied_units', 0) or 0)
        available_units = int(data.get('available_units', 0) or 0)
        open_requests = int(data.get('open_requests', 0) or 0)
        pending_tasks = int(data.get('pending_tasks', 0) or 0)
        avg_monthly_rent = float(data.get('avg_monthly_rent', 0) or 0)
        
        current_app.logger.debug(f'PDF Metrics - revenue={total_revenue}, units={total_units}, occupancy={occupancy_rate}%, open_requests={open_requests}, avg_rent={avg_monthly_rent}')
        
        metrics_data = [
            ['Metric', 'Value'],
            ['Total Revenue (MTD)', f"₱{total_revenue:,.2f}"],
            ['Outstanding Balance', f"₱{outstanding_balance:,.2f}"],
            ['Overdue Bills', str(overdue_bills)],
            ['Occupancy Rate', f"{occupancy_rate:.2f}%"],
            ['Total Units', str(total_units)],
            ['Occupied Units', str(occupied_units)],
            ['Available Units', str(available_units)],
            ['Open Requests', str(open_requests)],
            ['Pending Tasks', str(pending_tasks)],
            ['Avg Monthly Rent', f"₱{avg_monthly_rent:,.2f}"]
        ]
        
        metrics_table = Table(metrics_data, colWidths=[3*inch, 2*inch])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        story.append(Paragraph("Key Metrics", styles['Heading2']))
        story.append(Spacer(1, 0.1*inch))
        story.append(metrics_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Monthly Revenue
        if data['monthly_revenue']:
            story.append(Paragraph("Monthly Revenue", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            monthly_data_table = [['Month', 'Revenue']]
            for month in data['monthly_revenue']:
                monthly_data_table.append([
                    month.get('month', ''),
                    f"₱{float(month.get('revenue', 0)):,.2f}"
                ])
            
            monthly_table = Table(monthly_data_table, colWidths=[3*inch, 2*inch])
            monthly_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            story.append(monthly_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Occupancy by Unit Type
        if data['unit_type_breakdown']:
            story.append(Paragraph("Occupancy by Unit Type", styles['Heading2']))
            story.append(Spacer(1, 0.1*inch))
            occupancy_data_table = [['Unit Type', 'Occupancy %', 'Total', 'Occupied', 'Available']]
            for item in data['unit_type_breakdown']:
                occupancy_data_table.append([
                    item.get('type', 'All Units'),
                    f"{float(item.get('occupancy_rate', 0)):.2f}%",
                    str(item.get('total', 0)),
                    str(item.get('occupied', 0)),
                    str(item.get('available', 0))
                ])
            
            occupancy_table = Table(occupancy_data_table, colWidths=[2*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch])
            occupancy_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            story.append(occupancy_table)
        
        # Validate story has content
        if not story:
            current_app.logger.error('PDF story is empty - no content to generate')
            return jsonify({'error': 'No content to generate in PDF report'}), 400
        
        # Log story content for debugging
        current_app.logger.debug(f'PDF story has {len(story)} elements before build')
        
        # Build PDF
        try:
            doc.build(story)
            buffer.seek(0)
            
            # Verify buffer has content
            buffer_size = len(buffer.getvalue())
            current_app.logger.debug(f'PDF buffer size after build: {buffer_size} bytes')
            
            if buffer_size == 0:
                current_app.logger.error('PDF buffer is empty after build')
                return jsonify({'error': 'PDF generation failed - empty buffer'}), 500
            
            if buffer_size < 1000:  # Very small PDF might indicate an issue
                current_app.logger.warning(f'PDF buffer is very small: {buffer_size} bytes - might be empty content')
            
            # Generate filename
            filename = f"analytics_report_{datetime.now().strftime('%Y%m%d')}.pdf"
            
            response = make_response(send_file(
                buffer,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=filename
            ))
            response.headers['Content-Length'] = str(buffer_size)
            current_app.logger.debug(f'PDF report generated successfully: {filename}, {buffer_size} bytes')
            return response
        except Exception as build_error:
            current_app.logger.error(f'Error building PDF: {build_error}', exc_info=True)
            import traceback
            current_app.logger.error(f'PDF build traceback: {traceback.format_exc()}')
            return jsonify({'error': f'Failed to build PDF: {str(build_error)}'}), 500
        
    except Exception as e:
        current_app.logger.error(f'PDF report generation error: {e}', exc_info=True)
        return jsonify({'error': 'Failed to generate PDF report'}), 500


@analytics_bp.route('/download/excel', methods=['GET'])
@jwt_required()
def download_excel_report():
    """
    Download Excel report
    ---
    tags:
      - Analytics
    summary: Download analytics report as Excel
    description: Download analytics report as Excel. Property Manager only.
    security:
      - Bearer: []
    parameters:
      - in: query
        name: report_type
        type: string
        enum: [dashboard, financial, occupancy]
        default: dashboard
      - in: query
        name: start_date
        type: string
        format: date
      - in: query
        name: end_date
        type: string
        format: date
    responses:
      200:
        description: Excel report file
        schema:
          type: file
      401:
        description: Unauthorized
      403:
        description: Forbidden - Property Manager access required
      500:
        description: Server error
    """
    if not OPENPYXL_AVAILABLE:
        return jsonify({'error': 'Excel generation not available. Please install openpyxl.'}), 503
    
    try:
        property_id = get_property_id_from_request()
        if not property_id:
            return jsonify({'error': 'Property ID is required'}), 400
        
        data = _get_analytics_data_for_report(property_id)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"
        
        # Styles
        header_fill = PatternFill(start_color="1a1a1a", end_color="1a1a1a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title
        ws.merge_cells('A1:B1')
        ws['A1'] = "Property Analytics Report"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        
        # Report info
        ws['A2'] = f"Property: {data['property_name']}"
        ws['A3'] = f"Generated: {data['generated_at']}"
        
        row = 5
        
        # Key Metrics
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        metrics_headers = ['Metric', 'Value']
        for col, header in enumerate(metrics_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        row += 1
        # Safely get values with defaults for Excel
        total_revenue = float(data.get('total_revenue', 0) or 0)
        outstanding_balance = float(data.get('outstanding_balance', 0) or 0)
        overdue_bills = int(data.get('overdue_bills', 0) or 0)
        occupancy_rate = float(data.get('occupancy_rate', 0) or 0)
        total_units = int(data.get('total_units', 0) or 0)
        occupied_units = int(data.get('occupied_units', 0) or 0)
        available_units = int(data.get('available_units', 0) or 0)
        open_requests = int(data.get('open_requests', 0) or 0)
        pending_tasks = int(data.get('pending_tasks', 0) or 0)
        avg_monthly_rent = float(data.get('avg_monthly_rent', 0) or 0)
        
        metrics_data = [
            ['Total Revenue (MTD)', f"₱{total_revenue:,.2f}"],
            ['Outstanding Balance', f"₱{outstanding_balance:,.2f}"],
            ['Overdue Bills', str(overdue_bills)],
            ['Occupancy Rate', f"{occupancy_rate:.2f}%"],
            ['Total Units', str(total_units)],
            ['Occupied Units', str(occupied_units)],
            ['Available Units', str(available_units)],
            ['Open Requests', str(open_requests)],
            ['Pending Tasks', str(pending_tasks)],
            ['Avg Monthly Rent', f"₱{avg_monthly_rent:,.2f}"]
        ]
        
        for metric_row in metrics_data:
            for col, value in enumerate(metric_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
            row += 1
        
        row += 2
        
        # Monthly Revenue
        if data['monthly_revenue']:
            ws[f'A{row}'] = "Monthly Revenue"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            monthly_headers = ['Month', 'Revenue']
            for col, header in enumerate(monthly_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            for month in data['monthly_revenue']:
                ws.cell(row=row, column=1).value = month.get('month', '')
                ws.cell(row=row, column=2).value = f"₱{float(month.get('revenue', 0)):,.2f}"
                for col in range(1, 3):
                    ws.cell(row=row, column=col).border = border
                row += 1
            
            row += 2
        
        # Occupancy by Unit Type
        if data['unit_type_breakdown']:
            ws[f'A{row}'] = "Occupancy by Unit Type"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            occupancy_headers = ['Unit Type', 'Occupancy %', 'Total', 'Occupied', 'Available']
            for col, header in enumerate(occupancy_headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            for item in data['unit_type_breakdown']:
                ws.cell(row=row, column=1).value = item.get('type', 'All Units')
                ws.cell(row=row, column=2).value = f"{float(item.get('occupancy_rate', 0)):.2f}%"
                ws.cell(row=row, column=3).value = item.get('total', 0)
                ws.cell(row=row, column=4).value = item.get('occupied', 0)
                ws.cell(row=row, column=5).value = item.get('available', 0)
                for col in range(1, 6):
                    ws.cell(row=row, column=col).border = border
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generate filename
        filename = f"analytics_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f'Excel report generation error: {e}', exc_info=True)
        return jsonify({'error': 'Failed to generate Excel report'}), 500


@analytics_bp.route('/download/csv', methods=['GET'])
@jwt_required()
def download_csv_report():
    """
    Download CSV report
    ---
    tags:
      - Analytics
    summary: Download analytics report as CSV
    description: Download analytics report as CSV. Property Manager only.
    security:
      - Bearer: []
    parameters:
      - in: query
        name: report_type
        type: string
        enum: [dashboard, financial, occupancy]
        default: dashboard
      - in: query
        name: start_date
        type: string
        format: date
      - in: query
        name: end_date
        type: string
        format: date
    responses:
      200:
        description: CSV report file
        schema:
          type: file
      401:
        description: Unauthorized
      403:
        description: Forbidden - Property Manager access required
      500:
        description: Server error
    """
    try:
        property_id = get_property_id_from_request()
        if not property_id:
            return jsonify({'error': 'Property ID is required'}), 400
        
        data = _get_analytics_data_for_report(property_id)
        
        # Create CSV in memory
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        # Header
        writer.writerow(["Property Analytics Report"])
        writer.writerow([f"Property: {data['property_name']}"])
        writer.writerow([f"Generated: {data['generated_at']}"])
        writer.writerow([])
        
        # Key Metrics - safely get values with defaults
        total_revenue = float(data.get('total_revenue', 0) or 0)
        outstanding_balance = float(data.get('outstanding_balance', 0) or 0)
        overdue_bills = int(data.get('overdue_bills', 0) or 0)
        occupancy_rate = float(data.get('occupancy_rate', 0) or 0)
        total_units = int(data.get('total_units', 0) or 0)
        occupied_units = int(data.get('occupied_units', 0) or 0)
        available_units = int(data.get('available_units', 0) or 0)
        open_requests = int(data.get('open_requests', 0) or 0)
        pending_tasks = int(data.get('pending_tasks', 0) or 0)
        avg_monthly_rent = float(data.get('avg_monthly_rent', 0) or 0)
        
        writer.writerow(["Key Metrics"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Revenue (MTD)", f"₱{total_revenue:,.2f}"])
        writer.writerow(["Outstanding Balance", f"₱{outstanding_balance:,.2f}"])
        writer.writerow(["Overdue Bills", str(overdue_bills)])
        writer.writerow(["Occupancy Rate", f"{occupancy_rate:.2f}%"])
        writer.writerow(["Total Units", str(total_units)])
        writer.writerow(["Occupied Units", str(occupied_units)])
        writer.writerow(["Available Units", str(available_units)])
        writer.writerow(["Open Requests", str(open_requests)])
        writer.writerow(["Pending Tasks", str(pending_tasks)])
        writer.writerow(["Avg Monthly Rent", f"₱{avg_monthly_rent:,.2f}"])
        writer.writerow([])
        
        # Monthly Revenue
        if data['monthly_revenue']:
            writer.writerow(["Monthly Revenue"])
            writer.writerow(["Month", "Revenue"])
            for month in data['monthly_revenue']:
                writer.writerow([
                    month.get('month', ''),
                    f"₱{float(month.get('revenue', 0)):,.2f}"
                ])
            writer.writerow([])
        
        # Occupancy by Unit Type
        if data['unit_type_breakdown']:
            writer.writerow(["Occupancy by Unit Type"])
            writer.writerow(["Unit Type", "Occupancy %", "Total", "Occupied", "Available"])
            for item in data['unit_type_breakdown']:
                writer.writerow([
                    item.get('type', 'All Units'),
                    f"{float(item.get('occupancy_rate', 0)):.2f}%",
                    str(item.get('total', 0)),
                    str(item.get('occupied', 0)),
                    str(item.get('available', 0))
                ])
        
        # Convert to bytes
        csv_bytes = buffer.getvalue().encode('utf-8-sig')
        buffer.close()
        
        # Create response
        response = make_response(csv_bytes)
        filename = f"analytics_report_{datetime.now().strftime('%Y%m%d')}.csv"
        response.headers['Content-Type'] = 'text/csv; charset=utf-8-sig'
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        current_app.logger.error(f'CSV report generation error: {e}', exc_info=True)
        return jsonify({'error': 'Failed to generate CSV report'}), 500
